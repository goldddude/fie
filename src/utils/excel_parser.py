"""
Excel/CSV file parsing utilities
"""
import pandas as pd
from openpyxl import load_workbook


def parse_student_file(file_path, file_type='excel'):
    """
    Parse Excel or CSV file containing student data
    
    Args:
        file_path: Path to the file
        file_type: 'excel' or 'csv'
        
    Returns:
        tuple: (success, data_or_error)
            - If success: (True, list of student dictionaries)
            - If error: (False, error_message)
    """
    try:
        # Read file based on type
        if file_type == 'excel':
            df = pd.read_excel(file_path)
        else:
            df = pd.read_csv(file_path)
        
        # Expected columns
        required_columns = ['Name', 'Register Number', 'Section', 'Department', 'Duration']
        
        # Check if all required columns exist (case-insensitive)
        df.columns = df.columns.str.strip()
        missing_columns = []
        
        for col in required_columns:
            if col not in df.columns:
                # Try case-insensitive match
                found = False
                for df_col in df.columns:
                    if df_col.lower() == col.lower():
                        df.rename(columns={df_col: col}, inplace=True)
                        found = True
                        break
                if not found:
                    missing_columns.append(col)
        
        if missing_columns:
            return False, f"Missing required columns: {', '.join(missing_columns)}"
        
        # Convert to list of dictionaries
        students = []
        for idx, row in df.iterrows():
            # Skip empty rows
            if pd.isna(row['Name']) or pd.isna(row['Register Number']):
                continue
            
            student = {
                'name': str(row['Name']).strip(),
                'register_number': str(row['Register Number']).strip(),
                'section': str(row['Section']).strip(),
                'department': str(row['Department']).strip(),
                'duration': str(row['Duration']).strip()
            }
            students.append(student)
        
        if not students:
            return False, "No valid student data found in file"
        
        return True, students
        
    except Exception as e:
        return False, f"Error parsing file: {str(e)}"


def create_sample_excel(output_path):
    """
    Create a sample Excel template for student data
    
    Args:
        output_path: Path where the template should be saved
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    
    # Headers
    headers = ['Name', 'Register Number', 'Section', 'Department', 'Duration']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sample data
    sample_data = [
        ['John Doe', '2021CS001', 'A', 'Computer Science', 'Year 3'],
        ['Jane Smith', '2021CS002', 'A', 'Computer Science', 'Year 3'],
        ['Alice Johnson', '2021EC001', 'B', 'Electronics', 'Year 2'],
    ]
    
    for row_data in sample_data:
        ws.append(row_data)
    
    # Adjust column widths
    column_widths = [25, 20, 12, 25, 15]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    
    # Save
    wb.save(output_path)
    print(f"✅ Sample template created: {output_path}")
