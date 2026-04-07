"""
Working run script - MongoDB backend
"""
import os
import sys
from dotenv import load_dotenv
load_dotenv()

# Set environment
os.environ.setdefault('FLASK_ENV', 'development')
os.environ.setdefault('SECRET_KEY', 'dev-secret-key-change-in-production')

from flask import Flask, send_from_directory, request, jsonify
from flask_cors import CORS

# Create Flask app
app = Flask(__name__, static_folder='src/static', static_url_path='')

# Configuration
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'dev-secret-key-change-in-production')

# Initialize CORS
CORS(app)

# Initialize MongoDB connection
from src.database import get_db
with app.app_context():
    try:
        db = get_db()
        print("✅ Connected to MongoDB successfully")
    except Exception as e:
        print(f"❌ MongoDB connection failed: {e}")
        print("   Make sure MONGODB_URI is set in your .env file")
        sys.exit(1)


# ========== FRONTEND ROUTES ==========

@app.route('/')
def index():
    return send_from_directory('src/static', 'index.html')

@app.route('/<path:path>')
def serve_static(path):
    try:
        return send_from_directory('src/static', path)
    except:
        return send_from_directory('src/static', 'index.html')

# Serve Excel template
@app.route('/sample_students_template.xlsx')
def download_template():
    from flask import send_file
    template_path = os.path.join(os.path.dirname(__file__), 'sample_students_template.xlsx')
    if os.path.exists(template_path):
        return send_file(template_path, as_attachment=True, download_name='sample_students_template.xlsx')
    return {'error': 'Template not found'}, 404

# ========== STUDENT API ==========

@app.route('/api/students', methods=['GET', 'POST'])
def students_route():
    from src.services.student_service import StudentService
    if request.method == 'POST':
        data = request.get_json() or {}
        success, result = StudentService.create_student(data)
        if success:
            return {'message': 'Student created successfully', 'student': result.to_dict()}, 201
        return {'error': result}, 400
    else:
        search = request.args.get('search')
        filters = {}
        if request.args.get('section'):
            filters['section'] = request.args.get('section')
        if request.args.get('has_nfc'):
            filters['has_nfc'] = request.args.get('has_nfc').lower() == 'true'

        if search:
            students = StudentService.search_students(search)
        else:
            students = StudentService.get_all_students(filters)
        return {'count': len(students), 'students': [s.to_dict() for s in students]}, 200


@app.route('/api/students/<student_id>', methods=['GET', 'DELETE'])
def get_student_route(student_id):
    from src.services.student_service import StudentService
    student = StudentService.get_student_by_id(student_id)
    if not student:
        return {'error': 'Student not found'}, 404
    if request.method == 'DELETE':
        success, msg = StudentService.delete_student(student_id)
        return ({'message': msg} if success else {'error': msg}), (200 if success else 400)
    return {'student': student.to_dict()}, 200


@app.route('/api/students/upload', methods=['POST'])
def upload_students_route():
    if 'file' not in request.files:
        return {'error': 'No file provided'}, 400
    file = request.files['file']
    if file.filename == '':
        return {'error': 'No file selected'}, 400
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        return {'error': 'Invalid file format. Only .xlsx, .xls, and .csv are supported'}, 400

    try:
        from io import BytesIO, StringIO
        import csv
        from src.services.student_service import StudentService
        from datetime import datetime

        required_columns = ['name', 'register_number', 'section', 'department', 'duration']
        rows, headers = [], []

        if file.filename.endswith('.csv'):
            content = file.read().decode('utf-8-sig')
            csv_reader = csv.DictReader(StringIO(content))
            headers = [h.strip().lower().replace(' ', '_') for h in csv_reader.fieldnames]
            rows = list(csv_reader)
        else:
            from openpyxl import load_workbook
            wb = load_workbook(filename=BytesIO(file.read()), read_only=True)
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(h).strip().lower().replace(' ', '_') for h in header_row if h]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell for cell in row):
                    rows.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})

        missing = [c for c in required_columns if c not in headers]
        if missing:
            return {'error': f'Missing required columns: {", ".join(missing)}'}, 400

        success_count, failed_count, errors = 0, 0, []
        for idx, row in enumerate(rows):
            name = row.get('name')
            reg  = row.get('register_number')
            if not name or not reg:
                errors.append({'row': idx+2, 'register_number': str(reg or 'N/A'), 'error': 'Missing name or register number'})
                failed_count += 1
                continue
            data = {
                'name': str(name).strip(),
                'register_number': str(reg).strip(),
                'section': str(row.get('section', '')).strip(),
                'department': str(row.get('department', '')).strip(),
                'duration': str(row.get('duration', '')).strip(),
            }
            ok, res = StudentService.create_student(data)
            if ok:
                success_count += 1
            else:
                failed_count += 1
                errors.append({'row': idx+2, 'register_number': data['register_number'], 'error': res})

        return {'message': f'Upload complete: {success_count} added, {failed_count} failed',
                'success_count': success_count, 'failed_count': failed_count, 'errors': errors}, 200

    except Exception as e:
        return {'error': f'Failed to process file: {str(e)}'}, 500


# ========== NFC API ==========

@app.route('/api/nfc/register', methods=['POST'])
def register_nfc_route():
    from src.services.nfc_service import NFCService
    data = request.get_json() or {}
    student_id = data.get('student_id')
    nfc_tag_id = data.get('nfc_tag_id')
    if not student_id or not nfc_tag_id:
        return {'error': 'student_id and nfc_tag_id are required'}, 400
    success, result = NFCService.register_tag(student_id, nfc_tag_id)
    if success:
        return {'message': 'NFC tag registered successfully', 'student': result.to_dict()}, 200
    return {'error': result}, 400


@app.route('/api/nfc/student/<nfc_tag_id>', methods=['GET'])
def get_student_by_tag_route(nfc_tag_id):
    from src.services.nfc_service import NFCService
    student = NFCService.get_student_by_tag(nfc_tag_id)
    if not student:
        return {'error': 'No student found with this NFC tag'}, 404
    return {'student': student.to_dict()}, 200


# ========== FACULTY API ==========

@app.route('/api/faculty/login', methods=['POST'])
def faculty_login_route():
    from src.services.faculty_service import FacultyService
    data = request.get_json() or {}
    email = data.get('email')
    name  = data.get('name')
    if not email:
        return {'error': 'Email is required'}, 400
    success, result = FacultyService.generate_otp(email, name)
    if success:
        return {'message': 'OTP sent', 'email': email, 'otp': result}, 200
    return {'error': result}, 400


@app.route('/api/faculty/verify-otp', methods=['POST'])
def verify_otp_route():
    from src.services.faculty_service import FacultyService
    data = request.get_json() or {}
    email = data.get('email')
    otp   = data.get('otp')
    remember_me = data.get('remember_me', False)
    if not email or not otp:
        return {'error': 'Email and OTP are required'}, 400
    success, result = FacultyService.verify_otp(email, otp, remember_me)
    if not success:
        return {'error': result}, 400
    faculty, token = result
    resp = {'message': 'Login successful', 'faculty': faculty.to_dict()}
    if token:
        resp['remember_token'] = token
    return resp, 200


@app.route('/api/faculty/verify-token', methods=['POST'])
def verify_token_route():
    from src.services.faculty_service import FacultyService
    data = request.get_json() or {}
    email = data.get('email')
    token = data.get('token')
    if not email or not token:
        return {'error': 'Email and token are required'}, 400
    success, result = FacultyService.verify_remember_token(email, token)
    if not success:
        return {'error': result}, 400
    return {'message': 'Token valid', 'faculty': result.to_dict()}, 200


@app.route('/api/faculty/logout', methods=['POST'])
def faculty_logout_route():
    from src.services.faculty_service import FacultyService
    data = request.get_json() or {}
    email = data.get('email')
    if not email:
        return {'error': 'Email is required'}, 400
    FacultyService.logout(email)
    return {'message': 'Logged out successfully'}, 200


@app.route('/api/faculty/profile', methods=['GET'])
def get_faculty_profile_route():
    from src.services.faculty_service import FacultyService
    email = request.args.get('email')
    if not email:
        return {'error': 'Email is required'}, 400
    faculty = FacultyService.get_faculty_by_email(email)
    if faculty:
        return faculty.to_dict(), 200
    return {'error': 'Faculty not found'}, 404


# ========== ATTENDANCE API ==========

@app.route('/api/attendance/record', methods=['POST'])
def record_attendance_route():
    from src.services.attendance_service import AttendanceService
    from src.services.nfc_service import NFCService
    data = request.get_json() or {}
    student_id   = data.get('student_id')
    nfc_tag_id   = data.get('nfc_tag_id')
    faculty_name = data.get('faculty_name', 'Unknown Faculty')
    section      = data.get('section')
    subject      = data.get('subject')
    date         = data.get('date') or data.get('currentDate')
    class_time   = data.get('time') or data.get('currentTime')

    if nfc_tag_id and not student_id:
        stu = NFCService.get_student_by_tag(nfc_tag_id)
        if not stu:
            return {'error': 'No student found with this NFC tag'}, 404
        student_id = stu.id

    if not student_id:
        return {'error': 'student_id or nfc_tag_id is required'}, 400

    success, result = AttendanceService.record_attendance(
        student_id, faculty_name,
        section=section, subject=subject,
        date=date, class_time=class_time
    )
    if success:
        return {'message': 'Attendance recorded successfully', 'attendance': result.to_dict()}, 201
    return {'error': result}, 400


@app.route('/api/attendance/student/<student_id>', methods=['GET'])
def get_student_attendance_route(student_id):
    from src.services.attendance_service import AttendanceService
    limit = request.args.get('limit', type=int)
    records = AttendanceService.get_attendance_by_student(student_id, limit)
    return {'count': len(records), 'attendance': [r.to_dict() for r in records]}, 200


@app.route('/api/attendance/recent', methods=['GET'])
def get_recent_attendance_route():
    from src.services.attendance_service import AttendanceService
    limit = request.args.get('limit', 50, type=int)
    records = AttendanceService.get_recent_attendance(limit)
    return {'count': len(records), 'attendance': [r.to_dict() for r in records]}, 200


@app.route('/api/attendance/stats', methods=['GET'])
def get_stats_route():
    from src.services.attendance_service import AttendanceService
    return AttendanceService.get_attendance_stats(), 200


@app.route('/api/attendance/section-stats', methods=['GET'])
def get_section_stats_route():
    from src.services.attendance_service import AttendanceService
    return {'sections': AttendanceService.get_section_attendance_stats()}, 200


@app.route('/api/attendance/by-session', methods=['GET'])
def get_attendance_by_session_route():
    from src.database import get_db
    mongo = get_db()
    query = {}
    faculty = request.args.get('faculty')
    subject = request.args.get('subject')
    date    = request.args.get('date')
    section = request.args.get('section')
    if faculty:  query['recorded_by'] = faculty
    if subject:  query['subject']     = subject
    if date:     query['date']        = date
    if section:  query['section']     = section

    docs = list(mongo.attendance.find(query).sort('timestamp', -1))
    result = []
    for d in docs:
        result.append({
            'id': str(d.get('_id', '')),
            'student_id': d.get('student_id', ''),
            'student_name': d.get('student_name', ''),
            'register_number': d.get('register_number', ''),
            'recorded_by': d.get('recorded_by', ''),
            'section': d.get('section', ''),
            'subject': d.get('subject', ''),
            'date': d.get('date', ''),
            'class_time': d.get('class_time', ''),
            'timestamp': d['timestamp'].isoformat() if d.get('timestamp') else '',
            'session_closed': d.get('session_closed', False),
        })
    return {'count': len(result), 'attendance': result}, 200


@app.route('/api/attendance/close-session', methods=['POST'])
def close_session_route():
    from src.database import get_db
    mongo = get_db()
    data = request.get_json() or {}
    faculty_name = data.get('faculty_name')
    subject      = data.get('subject')
    date         = data.get('date')
    if not all([faculty_name, subject, date]):
        return {'error': 'faculty_name, subject, and date are required'}, 400
    result = mongo.attendance.update_many(
        {'recorded_by': faculty_name, 'subject': subject, 'date': date},
        {'$set': {'session_closed': True}}
    )
    if result.matched_count == 0:
        return {'error': 'No session records found'}, 404
    return {'message': 'Session closed successfully', 'closed_count': result.modified_count}, 200


@app.route('/api/attendance/<attendance_id>', methods=['DELETE'])
def delete_attendance_route(attendance_id):
    from src.database import get_db, to_object_id
    mongo = get_db()
    oid = to_object_id(attendance_id)
    result = mongo.attendance.delete_one({'_id': oid})
    if result.deleted_count == 0:
        return {'error': 'Attendance record not found'}, 404
    return {'message': 'Attendance record deleted successfully'}, 200


# Error handlers
@app.errorhandler(404)
def not_found(e):
    return {'error': 'Resource not found'}, 404

@app.errorhandler(500)
def server_error(e):
    return {'error': 'Internal server error'}, 500
    if request.method == 'POST':
        data = request.get_json()
        
        # Validate required fields
        required = ['name', 'register_number', 'section', 'department', 'duration']
        for field in required:
            if field not in data:
                return {'error': f'Missing field: {field}'}, 400
        
        # Check for duplicate register number
        existing = Student.query.filter_by(register_number=data['register_number'].strip()).first()
        if existing:
            return {'error': f"Student with register number {data['register_number']} already exists"}, 400
        
        # Create student
        student = Student(
            name=data['name'].strip(),
            register_number=data['register_number'].strip(),
            section=data['section'].strip(),
            department=data['department'].strip(),
            duration=data['duration'].strip()
        )
        
        db.session.add(student)
        db.session.commit()
        
        return {'message': 'Student created successfully', 'student': student.to_dict()}, 201
    
    else:  # GET
        search = request.args.get('search')
        section = request.args.get('section')
        has_nfc = request.args.get('has_nfc')
        
        query = Student.query
        
        if search:
            query = query.filter(db.or_(
                Student.name.ilike(f'%{search}%'),
                Student.register_number.ilike(f'%{search}%')
            ))
        
        if section:
            query = query.filter_by(section=section)
        
        if has_nfc:
            if has_nfc.lower() == 'true':
                query = query.filter(Student.nfc_tag_id.isnot(None))
            else:
                query = query.filter(Student.nfc_tag_id.is_(None))
        
        students = query.order_by(Student.register_number).all()
        
        return {
            'count': len(students),
            'students': [s.to_dict() for s in students]
        }, 200

@app.route('/api/students/<int:student_id>', methods=['GET', 'DELETE'])
def get_student(student_id):
    student = Student.query.get(student_id)
    if not student:
        return {'error': 'Student not found'}, 404
    
    if request.method == 'DELETE':
        # Delete student and all their attendance records (cascade)
        db.session.delete(student)
        db.session.commit()
        return {'message': 'Student deleted successfully'}, 200
    
    return {'student': student.to_dict()}, 200

# Upload students from Excel/CSV
@app.route('/api/students/upload', methods=['POST'])
def upload_students():
    if 'file' not in request.files:
        return {'error': 'No file provided'}, 400
    
    file = request.files['file']
    
    if file.filename == '':
        return {'error': 'No file selected'}, 400
    
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        return {'error': 'Invalid file format. Only .xlsx, .xls, and .csv are supported'}, 400
    
    try:
        from io import BytesIO, StringIO
        import csv
        
        # Required columns matching student model
        required_columns = ['name', 'register_number', 'section', 'department', 'duration']
        
        rows = []
        headers = []
        
        # Read file based on extension
        if file.filename.endswith('.csv'):
            # CSV parsing
            content = file.read().decode('utf-8-sig')  # Handle BOM
            csv_reader = csv.DictReader(StringIO(content))
            headers = [h.strip().lower().replace(' ', '_') for h in csv_reader.fieldnames]
            rows = list(csv_reader)
        else:
            # Excel parsing using openpyxl
            try:
                from openpyxl import load_workbook
            except ImportError:
                return {'error': 'openpyxl not installed. Please install it: pip install openpyxl'}, 500
            
            wb = load_workbook(filename=BytesIO(file.read()), read_only=True)
            ws = wb.active
            
            # Get headers from first row
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(h).strip().lower().replace(' ', '_') for h in header_row if h]
            
            # Get data rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell for cell in row):  # Skip empty rows
                    row_dict = {}
                    for idx, header in enumerate(headers):
                        if idx < len(row):
                            row_dict[header] = row[idx]
                    rows.append(row_dict)
        
        # Check for required columns
        missing_columns = [col for col in required_columns if col not in headers]
        if missing_columns:
            return {'error': f'Missing required columns: {", ".join(missing_columns)}'}, 400
        
        success_count = 0
        failed_count = 0
        errors = []
        
        for index, row in enumerate(rows):
            try:
                # Get values with proper None handling
                name = row.get('name')
                register_num = row.get('register_number')
                section = row.get('section')
                department = row.get('department')
                duration = row.get('duration')
                
                # Validate required fields
                if not name or not register_num:
                    errors.append({
                        'row': index + 2,
                        'register_number': str(register_num) if register_num else 'N/A',
                        'error': 'Missing name or register number'
                    })
                    failed_count += 1
                    continue
                
                # Convert to string and strip
                register_num = str(register_num).strip()
                
                # Check if student already exists
                existing = Student.query.filter_by(register_number=register_num).first()
                if existing:
                    errors.append({
                        'row': index + 2,
                        'register_number': register_num,
                        'error': 'Student already exists'
                    })
                    failed_count += 1
                    continue
                
                # Create student with exact fields from model
                student = Student(
                    name=str(name).strip(),
                    register_number=register_num,
                    section=str(section).strip() if section else '',
                    department=str(department).strip() if department else '',
                    duration=str(duration).strip() if duration else ''
                )
                
                db.session.add(student)
                success_count += 1
                
            except Exception as e:
                errors.append({
                    'row': index + 2,
                    'register_number': str(row.get('register_number', 'N/A')),
                    'error': str(e)
                })
                failed_count += 1
        
        # Commit all successful students
        if success_count > 0:
            db.session.commit()
        
        return {
            'message': f'Upload complete: {success_count} students added, {failed_count} failed',
            'success_count': success_count,
            'failed_count': failed_count,
            'errors': errors
        }, 200
        
    except Exception as e:
        return {'error': f'Failed to process file: {str(e)}'}, 500



# ========== NFC API ==========

@app.route('/api/nfc/register', methods=['POST'])
def register_nfc():
    data = request.get_json()
    
    student_id = data.get('student_id')
    nfc_tag_id = data.get('nfc_tag_id')
    
    if not student_id or not nfc_tag_id:
        return {'error': 'student_id and nfc_tag_id are required'}, 400
    
    student = Student.query.get(student_id)
    if not student:
        return {'error': 'Student not found'}, 404
    
    # Check if tag already registered to another student
    existing = Student.query.filter_by(nfc_tag_id=nfc_tag_id.strip()).first()
    if existing and existing.id != student_id:
        return {'error': f'NFC tag already registered to {existing.name}'}, 400
    
    student.nfc_tag_id = nfc_tag_id.strip()
    db.session.commit()
    
    return {'message': 'NFC tag registered successfully', 'student': student.to_dict()}, 200

@app.route('/api/nfc/student/<nfc_tag_id>', methods=['GET'])
def get_student_by_tag(nfc_tag_id):
    student = Student.query.filter_by(nfc_tag_id=nfc_tag_id).first()
    if not student:
        return {'error': 'No student found with this NFC tag'}, 404
    return {'student': student.to_dict()}, 200

# ========== FACULTY API ==========

@app.route('/api/faculty/login', methods=['POST'])
def faculty_login():
    from src.models import Faculty
    import random
    from datetime import datetime
    
    data = request.get_json()
    
    if not data:
        return {'error': 'No data provided'}, 400
    
    email = data.get('email')
    name = data.get('name')
    
    if not email:
        return {'error': 'Email is required'}, 400
    
    # Generate 6-digit OTP
    otp = ''.join([str(random.randint(0, 9)) for _ in range(6)])
    
    # Find or create faculty
    faculty = Faculty.query.filter_by(email=email).first()
    
    if not faculty:
        if not name:
            return {'error': 'Name is required for new faculty'}, 400
        
        faculty = Faculty(
            name=name,
            email=email
        )
        db.session.add(faculty)
    
    # Update OTP
    faculty.otp = otp
    faculty.otp_created_at = datetime.utcnow()
    
    db.session.commit()
    
    print(f"📧 OTP for {email}: {otp}")
    
    return {
        'message': 'OTP sent to your email',
        'email': email,
        'otp': otp
    }, 200

@app.route('/api/faculty/verify-otp', methods=['POST'])
def verify_otp():
    from src.models import Faculty
    from datetime import datetime, timedelta
    import secrets
    
    data = request.get_json()
    
    if not data:
        return {'error': 'No data provided'}, 400
    
    email = data.get('email')
    otp = data.get('otp')
    remember_me = data.get('remember_me', False)
    
    if not email or not otp:
        return {'error': 'Email and OTP are required'}, 400
    
    faculty = Faculty.query.filter_by(email=email).first()
    
    if not faculty:
        return {'error': 'Faculty not found'}, 400
    
    if not faculty.otp:
        return {'error': 'No OTP generated. Please request a new one'}, 400
    
    # Check if OTP is expired (10 minutes)
    if faculty.otp_created_at:
        otp_age = datetime.utcnow() - faculty.otp_created_at
        if otp_age > timedelta(minutes=10):
            return {'error': 'OTP expired. Please request a new one'}, 400
    
    # Verify OTP
    if faculty.otp != otp:
        return {'error': 'Invalid OTP'}, 400
    
    # Clear OTP
    faculty.otp = None
    faculty.otp_created_at = None
    
    # Generate remember token if requested
    token = None
    if remember_me:
        token = secrets.token_urlsafe(32)
        faculty.remember_token = token
        faculty.remember_expires = datetime.utcnow() + timedelta(days=30)
    
    db.session.commit()
    
    response_data = {
        'message': 'Login successful',
        'faculty': faculty.to_dict()
    }
    
    if token:
        response_data['remember_token'] = token
    
    return response_data, 200

@app.route('/api/faculty/verify-token', methods=['POST'])
def verify_token():
    from src.models import Faculty
    from datetime import datetime
    
    data = request.get_json()
    
    if not data:
        return {'error': 'No data provided'}, 400
    
    email = data.get('email')
    token = data.get('token')
    
    if not email or not token:
        return {'error': 'Email and token are required'}, 400
    
    faculty = Faculty.query.filter_by(email=email).first()
    
    if not faculty:
        return {'error': 'Faculty not found'}, 400
    
    if not faculty.remember_token or faculty.remember_token != token:
        return {'error': 'Invalid token'}, 400
    
    # Check if token is expired
    if faculty.remember_expires and faculty.remember_expires < datetime.utcnow():
        return {'error': 'Token expired. Please login again'}, 400
    
    return {
        'message': 'Token valid',
        'faculty': faculty.to_dict()
    }, 200

@app.route('/api/faculty/logout', methods=['POST'])
def faculty_logout():
    from src.models import Faculty
    
    data = request.get_json()
    
    if not data:
        return {'error': 'No data provided'}, 400
    
    email = data.get('email')
    
    if not email:
        return {'error': 'Email is required'}, 400
    
    faculty = Faculty.query.filter_by(email=email).first()
    
    if faculty:
        faculty.remember_token = None
        faculty.remember_expires = None
        db.session.commit()
    
    return {'message': 'Logged out successfully'}, 200

@app.route('/api/faculty/profile', methods=['GET'])
def get_faculty_profile():
    from src.models import Faculty
    
    email = request.args.get('email')
    
    if not email:
        return {'error': 'Email is required'}, 400
    
    faculty = Faculty.query.filter_by(email=email).first()
    
    if faculty:
        return faculty.to_dict(), 200
    else:
        return {'error': 'Faculty not found'}, 404

# ========== ATTENDANCE API ==========

@app.route('/api/attendance/record', methods=['POST'])
def record_attendance():
    from datetime import datetime, timedelta
    
    data = request.get_json()
    
    student_id = data.get('student_id')
    nfc_tag_id = data.get('nfc_tag_id')
    faculty_name = data.get('faculty_name', 'Unknown Faculty')
    section = data.get('section')
    subject = data.get('subject')
    date = data.get('date')  # Date of class
    class_time = data.get('time')  # Class time slot
    
    # Get student by NFC tag if provided
    if nfc_tag_id and not student_id:
        student = Student.query.filter_by(nfc_tag_id=nfc_tag_id).first()
        if not student:
            return {'error': 'No student found with this NFC tag'}, 404
        student_id = student.id
    
    if not student_id:
        return {'error': 'student_id or nfc_tag_id is required'}, 400
    
    student = Student.query.get(student_id)
    if not student:
        return {'error': 'Student not found'}, 404
    
    # Check for duplicate within 1 hour
    one_hour_ago = datetime.utcnow() - timedelta(hours=1)
    recent = Attendance.query.filter(
        Attendance.student_id == student_id,
        Attendance.timestamp >= one_hour_ago
    ).first()
    
    if recent:
        return {'error': f"Attendance already recorded for {student.name} at {recent.timestamp.strftime('%H:%M:%S')}"}, 400
    
    # Create attendance record
    attendance = Attendance(
        student_id=student_id,
        recorded_by=faculty_name,
        section=section,
        subject=subject,
        date=date,
        class_time=class_time
    )
    
    db.session.add(attendance)
    db.session.commit()
    
    return {'message': 'Attendance recorded successfully', 'attendance': attendance.to_dict()}, 201

@app.route('/api/attendance/student/<int:student_id>', methods=['GET'])
def get_student_attendance(student_id):
    limit = request.args.get('limit', type=int)
    query = Attendance.query.filter_by(student_id=student_id).order_by(Attendance.timestamp.desc())
    
    if limit:
        query = query.limit(limit)
    
    records = query.all()
    
    return {
        'count': len(records),
        'attendance': [r.to_dict() for r in records]
    }, 200

@app.route('/api/attendance/recent', methods=['GET'])
def get_recent_attendance():
    limit = request.args.get('limit', 50, type=int)
    records = Attendance.query.order_by(Attendance.timestamp.desc()).limit(limit).all()
    
    return {
        'count': len(records),
        'attendance': [r.to_dict() for r in records]
    }, 200

@app.route('/api/attendance/stats', methods=['GET'])
def get_stats():
    from datetime import datetime
    
    total_students = Student.query.count()
    total_records = Attendance.query.count()
    
    # Today's attendance
    today = datetime.utcnow().date()
    from datetime import datetime as dt
    start_of_day = dt.combine(today, dt.min.time())
    today_count = Attendance.query.filter(Attendance.timestamp >= start_of_day).count()
    
    # Unique students today
    today_students = db.session.query(Attendance.student_id).filter(
        Attendance.timestamp >= start_of_day
    ).distinct().count()
    
    return {
        'total_students': total_students,
        'total_attendance_records': total_records,
        'today_attendance_count': today_count,
        'today_unique_students': today_students,
        'today_percentage': round((today_students / total_students * 100) if total_students > 0 else 0, 2)
    }, 200

# Get attendance filtered by faculty / subject / date / section
@app.route('/api/attendance/by-session', methods=['GET'])
def get_attendance_by_session():
    faculty = request.args.get('faculty')
    subject = request.args.get('subject')
    date    = request.args.get('date')
    section = request.args.get('section')

    query = Attendance.query

    if faculty:
        query = query.filter(Attendance.recorded_by == faculty)
    if subject:
        query = query.filter(Attendance.subject == subject)
    if date:
        try:
            from datetime import datetime as dt
            attendance_date = dt.strptime(date, '%Y-%m-%d').date()
            query = query.filter(Attendance.date == attendance_date)
        except ValueError:
            pass
    if section:
        # Join with Students table to filter by section
        query = query.join(Student, Attendance.student_id == Student.id).filter(Student.section == section)

    records = query.order_by(Attendance.timestamp.desc()).all()

    # Attach section info from student
    result = []
    for r in records:
        d = r.to_dict()
        d['section'] = r.student.section if r.student else None
        result.append(d)

    return {'count': len(result), 'attendance': result}, 200


# Close attendance session
@app.route('/api/attendance/close-session', methods=['POST'])
def close_session():
    data = request.get_json()
    faculty_name = data.get('faculty_name')
    subject      = data.get('subject')
    date         = data.get('date')

    if not all([faculty_name, subject, date]):
        return {'error': 'faculty_name, subject, and date are required'}, 400

    try:
        # 'date' arrives as a string "YYYY-MM-DD" and is stored that way in the DB
        # Do NOT convert to a date object — compare as string directly
        records = Attendance.query.filter(
            Attendance.recorded_by == faculty_name,
            Attendance.subject     == subject,
            Attendance.date        == date,      # string-to-string comparison
        ).all()

        if not records:
            return {'error': 'No session records found for this faculty/subject/date'}, 404

        closed_count = 0
        for record in records:
            try:
                record.session_closed = True     # graceful if column missing on old DB
                closed_count += 1
            except Exception:
                closed_count += 1                # still count it
        db.session.commit()

        return {'message': 'Session closed successfully', 'closed_count': closed_count}, 200
    except Exception as e:
        return {'error': str(e)}, 500


# Delete individual attendance record (enables re-scan)
@app.route('/api/attendance/<int:attendance_id>', methods=['DELETE'])
def delete_attendance_record(attendance_id):
    attendance = Attendance.query.get(attendance_id)
    if not attendance:
        return {'error': 'Attendance record not found'}, 404
    db.session.delete(attendance)
    db.session.commit()
    return {'message': 'Attendance record deleted successfully'}, 200


# Error handlers
@app.errorhandler(404)
def not_found(e):
    return {'error': 'Resource not found'}, 404

@app.errorhandler(500)
def server_error(e):
    return {'error': 'Internal server error'}, 500

if __name__ == '__main__':
    import socket
    
    # Get local IP
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
        s.close()
    except:
        local_ip = "Unable to detect"
    
    host = '0.0.0.0'  # Listen on all interfaces
    port = 5000
    
    print(f"""
    ╔═══════════════════════════════════════════════════════╗
    ║     NFC Attendance System - Development Server        ║
    ╚═══════════════════════════════════════════════════════╝
    
    ✅ All APIs loaded successfully!
    
    🌐 Access from Computer: http://localhost:{port}
    📱 Access from Android:   http://{local_ip}:{port}
    
    🗄️  Database: SQLite (development)
    
    💡 Make sure your Android phone is on the same WiFi network!
    💡 For NFC features, use Chrome browser on Android
    
    Press CTRL+C to stop the server
    """)
    
    app.run(host=host, port=port, debug=True)
